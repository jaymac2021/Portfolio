from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests

# import workbook
wb = load_workbook("/Users/william/Downloads/Riipen Vehicle-Price Updated.xlsx")

# set active worksheet on workbook
ws = wb.active

# set the source of where the requests are coming from
source = requests.get("https://www.caa.ca/gas-prices/").text

# set variable for beautifulsoup
soup = BeautifulSoup(source, "lxml")

# set variable and find information from website
price = soup.find("div", class_="national_single_price").text

# clean information given from website
price = price.replace('/L', '')

# set information as a float number
canadaAverageGasPrice = float(price)

# set gas and electricity variables
canadaAverageElectricityPrice = ws["H14"].value

# set variables for ICE Ford Truck
ford = ws["B4"].value

fordModel = ws["C4"].value

fordPrice1 = ws["D4"].value
fordPrice2 = ws["D5"].value
fordPrice3 = ws["D6"].value

fordEngine1 = ws["E4"].value
fordEngine2 = ws["E5"].value
fordEngine3 = ws["E6"].value

fordFuelCap1 = ws["F4"].value
fordFuelCap2 = ws["F5"].value
fordFuelCap3 = ws["F6"].value

fordFuel1 = ws["G4"].value
fordFuel2 = ws["G5"].value
fordFuel3 = ws["G6"].value

fordCityMilage1 = ws["H4"].value
fordCityMilage2 = ws["H5"].value
fordCityMilage3 = ws["H6"].value

fordHighwayMilage1 = ws["I4"].value
fordHighwayMilage2 = ws["I5"].value
fordHighwayMilage3 = ws["I6"].value

fordCityVehicleRunningCost1 = canadaAverageGasPrice / fordCityMilage1
fordCityVehicleRunningCost2 = canadaAverageGasPrice / fordCityMilage2
fordCityVehicleRunningCost3 = canadaAverageGasPrice / fordCityMilage3

fordHighwayVehicleRunningCost1 = canadaAverageGasPrice / fordHighwayMilage1
fordHighwayVehicleRunningCost2 = canadaAverageGasPrice / fordHighwayMilage2
fordHighwayVehicleRunningCost3 = canadaAverageGasPrice / fordHighwayMilage3


# set variables for ICE GMC Truck
gmc = ws["B7"].value

gmcModel = ws["C7"].value

gmcPrice = ws["D7"].value

gmcEngine = ws["E7"].value

gmcFuelCap = ws["F7"].value

gmcFuel = ws["G7"].value

gmcCityMilage = ws["H7"].value

gmcHighwayMilage = ws["I7"].value

gmcCityVehicleRunningCost = canadaAverageGasPrice / gmcCityMilage

gmcHighwayVehicleRunningCost = canadaAverageGasPrice / gmcHighwayMilage


# set variables for ICE Chevrolet Truck
chev = ws["B8"].value

chevModel = ws["C8"].value

chevPrice = ws["D8"].value

chevEngine = ws["E8"].value

chevFuelCap = ws["F8"].value

chevFuel = ws["G8"].value

chevCityMilage = ws["H8"].value

chevHighwayMilage = ws["I8"].value

chevCityVehicleRunningCost = canadaAverageGasPrice / chevCityMilage

chevHighwayVehicleRunningCost = canadaAverageGasPrice / chevHighwayMilage




# set variables for EV Tesla Truck
tesla = ws["B14"].value

teslaModel = ws["C14"].value

teslaPrice1 = ws["D14"].value
teslaPrice2 = ws["D15"].value
teslaPrice3 = ws["D16"].value

teslaVariant1 = ws["E14"].value
teslaVariant2 = ws["E15"].value
teslaVariant3 = ws["E16"].value

teslaBatteryCap1 = ws["F14"].value
teslaBatteryCap2 = ws["F15"].value
teslaBatteryCap3 = ws["F16"].value

teslaElectricRange1 = ws["G14"].value
teslaElectricRange2 = ws["G15"].value
teslaElectricRange3 = ws["G16"].value

teslaCostToFullyCharge1 = (teslaBatteryCap1 * canadaAverageElectricityPrice) / 100
teslaCostToFullyCharge2 = (teslaBatteryCap2 * canadaAverageElectricityPrice) / 100
teslaCostToFullyCharge3 = (teslaBatteryCap3 * canadaAverageElectricityPrice) / 100

teslaVehicleRunningCost1 = (teslaCostToFullyCharge1 / teslaElectricRange1) * 100
teslaVehicleRunningCost2 = (teslaCostToFullyCharge2 / teslaElectricRange2) * 100
teslaVehicleRunningCost3 = (teslaCostToFullyCharge3 / teslaElectricRange3) * 100



# set variables for EV Ford Truck
evFord = ws["B17"].value

evFordModel = ws["C17"].value

evFordPrice1 = ws["D17"].value
evFordPrice2 = ws["D17"].value

evFordVariant1 = ws["E17"].value
evFordVariant2 = ws["E18"].value

evFordBatteryCap1 = ws["F17"].value
evFordBatteryCap2 = ws["F18"].value

evFordElectricRange1 = ws["G17"].value
evFordElectricRange2 = ws["G18"].value

evFordCostToFullyCharge1 = (evFordBatteryCap1 * canadaAverageElectricityPrice) / 100
evFordCostToFullyCharge2 = (evFordBatteryCap2 * canadaAverageElectricityPrice) / 100

evFordVehicleRunningCost1 = (evFordCostToFullyCharge1 / evFordElectricRange1) * 100
evFordVehicleRunningCost2 = (evFordCostToFullyCharge2 / evFordElectricRange2) * 100



# set variables for EV Rivian Truck
rivian = ws["B19"].value

rivianModel = ws["C19"].value

rivianPrice = ws["D19"].value

rivianVariant = ws["E19"].value

rivianBatteryCap = ws["F19"].value

rivianElectricRange = ws["G19"].value

rivianCostToFullyCharge = (rivianBatteryCap * canadaAverageElectricityPrice) / 100

rivianVehicleRunningCost = (rivianCostToFullyCharge / rivianElectricRange) * 100



# set variables for chosen ICE truck information
chosenICEMake = ""

chosenICEModel = ""

chosenICEPrice = 1

chosenICEEngine = ""

chosenICEFuelCap = 1

chosenICEFuel = ""

chosenICECityMilage = 1

chosenICEHighwayMilage = 1





# set variables for chosen EV truck information
chosenEVMake = ""

chosenEVModel = ""

chosenEVPrice = 1

chosenEVVariant = ""

chosenEVBatteryCap = 1

chosenEVElectricRange = 1



# begin application
print("")
print("Welcome to the ICE vs EV Comparison App")

# set application into loop
choosing = True
while (choosing):
    print("")
    print("*** Main Menu ***")
    print("1. List of ICE Trucks")
    print("2. List of EV Trucks")
    print("3. Compare EV TrucK to ICE Truck")
    print("4. Exit")
    print("****************")
    print("")

# assign variable that takes in user choice
    choice = input("Please make a selection from the Main Menu: ")
    print("")

# set actions based off user choice
    if (int(choice) == 4):
        print("Exiting App")
        print("")
        break

    elif (int(choice) == 1):
        print("1.",ford,fordModel)
        print("2.",gmc,gmcModel)
        print("3.",chev,chevModel)
        print("")
        iceTruck = input("Which ICE Truck would you like to view? ")
        print("")

        if (int(iceTruck) == 1):
            print("-----------------------------------------------------")
            print("Three different engines available to choose from, each engine information")
            print("will be separated by a |")
            print("For example: Engine 1 information | Engine 2 information | Engine 3 information")
            print("-----------------------------------------------------")
            print(ford,fordModel)
            print("")
            print("Price:",fordPrice1,"|",fordPrice2,"|",fordPrice3)
            print("")
            print("Engine:",fordEngine1,"|",fordEngine2,"|",fordEngine3)
            print("")
            print("Fuel Tank Capacity (Litres):",fordFuelCap1,"|",fordFuelCap2,"|",fordFuelCap3)
            print("")
            print("Fuel Type:",fordFuel1,"|",fordFuel2,"|",fordFuel3)
            print("")
            print("Milage (Km/L)")
            print("City:",f'{fordCityMilage1:.2f}',"|",f'{fordCityMilage2:.2f}',"|",f'{fordCityMilage3:.2f}')
            print("Highway:",f'{fordHighwayMilage1:.2f}',"|",f'{fordHighwayMilage2:.2f}',"|",f'{fordHighwayMilage3:.2f}')
            print("")
            print("Canada Average Gas Price (¢/L):",canadaAverageGasPrice)
            print("")
            print("Vehicle Running Cost (¢/km)")
            print("City:",f'{fordCityVehicleRunningCost1:.2f}',"|",f'{fordCityVehicleRunningCost2:.2f}',"|",f'{fordCityVehicleRunningCost3:.2f}')
            print("Highway:",f'{fordHighwayVehicleRunningCost1:.2f}',"|",f'{fordHighwayVehicleRunningCost2:.2f}',"|",f'{fordHighwayVehicleRunningCost3:.2f}')
            print("-----------------------------------------------------")
            input("Press enter to continue")
            print("")
        
        if (int(iceTruck) == 2):
            print("-----------------------------------------------------")
            print(gmc,gmcModel)
            print("")
            print("Price:",gmcPrice)
            print("")
            print("Engine:",gmcEngine)
            print("")
            print("Fuel Tank Capacity (Litres):",gmcFuelCap)
            print("")
            print("Fuel Type:",gmcFuel)
            print("")
            print("Milage (Km/L)")
            print("City:",f'{gmcCityMilage:.2f}')
            print("Highway:",f'{gmcHighwayMilage:.2f}')
            print("")
            print("Canada Average Gas Price (¢/L):",canadaAverageGasPrice)
            print("")
            print("Vehicle Running Cost (¢/km)")
            print("City:",f'{gmcCityVehicleRunningCost:.2f}')
            print("Highway:",f'{gmcHighwayVehicleRunningCost:.2f}')
            print("-----------------------------------------------------")
            input("Press enter to continue")
            print("")

        if (int(iceTruck) == 3):
            print("-----------------------------------------------------")
            print(chev,chevModel)
            print("")
            print("Price:",chevPrice)
            print("")
            print("Engine:",chevEngine)
            print("")
            print("Fuel Tank Capacity (Litres):",chevFuelCap)
            print("")
            print("Fuel Type:",chevFuel)
            print("")
            print("Milage (Km/L)")
            print("City:",f'{chevCityMilage:.2f}')
            print("Highway:",f'{chevHighwayMilage:.2f}')
            print("")
            print("Canada Average Gas Price (¢/L):",canadaAverageGasPrice)
            print("")
            print("Vehicle Running Cost (¢/km)")
            print("City:",f'{chevCityVehicleRunningCost:.2f}')
            print("Highway:",f'{chevHighwayVehicleRunningCost:.2f}')
            print("-----------------------------------------------------")
            input("Press enter to continue")
            print("")
        



    elif (int(choice) == 2):
        print("1.",tesla,teslaModel,"(To launch in 2023)")
        print("2.",ford,evFordModel)
        print("3.",rivian,rivianModel)
        print("")
        evTruck = input("Which EV Truck would you like to view? ")
        print("")

        if (int(evTruck) == 1):
            print("-----------------------------------------------------")
            print("Three different variants available to choose from, each variant information")
            print("will be separated by a |")
            print("For example: Variant 1 information | Variant 2 information | Variant 3 information")
            print("-----------------------------------------------------")
            print(tesla,teslaModel)
            print("")
            print("Price:",teslaPrice1,"|",teslaPrice2,"|",teslaPrice3)
            print("")
            print("Variant:",teslaVariant1,"|",teslaVariant2,"|",teslaVariant3)
            print("")
            print("Battery Capacity (KWh):",teslaBatteryCap1,"|",teslaBatteryCap2,"|",teslaBatteryCap3)
            print("")
            print("Electric Range (km):",teslaElectricRange1,"|",teslaElectricRange2,"|",teslaElectricRange3)
            print("")
            print("Canada Average Electricity Price (¢/kWh):",canadaAverageElectricityPrice)
            print("")
            print("Cost to Fully Charge EV ($):",teslaCostToFullyCharge1,"|",teslaCostToFullyCharge2,"|",teslaCostToFullyCharge3)
            print("")
            print("Vehicle Running Cost (¢/km)",f'{teslaVehicleRunningCost1:.2f}',"|",f'{teslaVehicleRunningCost2:.2f}',"|",f'{teslaVehicleRunningCost3:.2f}')
            print("-----------------------------------------------------")
            input("Press enter to continue")
            print("")

        if (int(evTruck) == 2):
            print("-----------------------------------------------------")
            print("Two different variants available to choose from, each variant information")
            print("will be separated by a |")
            print("For example: Variant 1 information | Variant 2 information")
            print("-----------------------------------------------------")
            print(evFord,evFordModel)
            print("")
            print("Price:",evFordPrice1,"|",evFordPrice2)
            print("")
            print("Variant:",evFordVariant1,"|",evFordVariant2)
            print("")
            print("Battery Capacity (KWh):",evFordBatteryCap1,"|",evFordBatteryCap2)
            print("")
            print("Electric Range (km):",evFordElectricRange1,"|",evFordElectricRange2)
            print("")
            print("Canada Average Electricity Price (¢/kWh):",canadaAverageElectricityPrice)
            print("")
            print("Cost to Fully Charge EV ($):",f'{evFordCostToFullyCharge1:.2f}',"|",f'{evFordCostToFullyCharge2:.2f}')
            print("")
            print("Vehicle Running Cost (¢/km):",f'{evFordVehicleRunningCost1:.2f}',"|",f'{evFordVehicleRunningCost2:.2f}')
            print("-----------------------------------------------------")
            input("Press enter to continue")
            print("")

        if (int(evTruck) == 3):
            print("-----------------------------------------------------")
            print(rivian,rivianModel)
            print("")
            print("Price:",rivianPrice)
            print("")
            print("Variant:",rivianVariant)
            print("")
            print("Battery Capacity (KWh):",rivianBatteryCap)
            print("")
            print("Electric Range (km):",rivianElectricRange)
            print("")
            print("Canada Average Electricity Price (¢/kWh):",canadaAverageElectricityPrice)
            print("")
            print("Cost to Fully Charge EV ($):",f'{rivianCostToFullyCharge:.2f}')
            print("")
            print("Vehicle Running Cost (¢/km):",f'{rivianVehicleRunningCost:.2f}')
            print("-----------------------------------------------------")
            input("Press enter to continue")
            print("")

    elif (int(choice) == 3):
        print("")
        print("1. Tesla",teslaModel)
        print("2. Ford",evFordModel)
        print("3. Rivian",rivianModel)
        print("")
        compareChoice2 = input("Pick an EV Truck to compare: ")
        print("")

        if (int(compareChoice2) == 1):
            print("Which variant for the Tesla",teslaModel)
            print("1.",teslaVariant1)
            print("2.",teslaVariant2)
            print("3.",teslaVariant3)
            print("")
            variantChoice = input("")
            print("")

            if (int(variantChoice) == 1):
                chosenEVMake = tesla

                chosenEVModel = teslaModel

                chosenEVPrice = teslaPrice1

                chosenEVVariant = teslaVariant1

                chosenEVBatteryCap = teslaBatteryCap1

                chosenEVElectricRange = teslaElectricRange1

                chosenEVCostToFullyCharge = (chosenEVBatteryCap * canadaAverageElectricityPrice) / 100

                chosenEVVehicleRunningCost = (chosenEVCostToFullyCharge / chosenEVElectricRange) * 100
            
            if (int(variantChoice) == 2):
                chosenEVMake = tesla

                chosenEVModel = teslaModel

                chosenEVPrice = teslaPrice2

                chosenEVVariant = teslaVariant2

                chosenEVBatteryCap = teslaBatteryCap2

                chosenEVElectricRange = teslaElectricRange2

                chosenEVCostToFullyCharge = (chosenEVBatteryCap * canadaAverageElectricityPrice) / 100

                chosenEVVehicleRunningCost = (chosenEVCostToFullyCharge / chosenEVElectricRange) * 100
            
            if (int(variantChoice) == 3):
                chosenEVMake = tesla

                chosenEVModel = teslaModel

                chosenEVPrice = teslaPrice3

                chosenEVVariant = teslaVariant3

                chosenEVBatteryCap = teslaBatteryCap3

                chosenEVElectricRange = teslaElectricRange3

                chosenEVCostToFullyCharge = (chosenEVBatteryCap * canadaAverageElectricityPrice) / 100

                chosenEVVehicleRunningCost = (chosenEVCostToFullyCharge / chosenEVElectricRange) * 100
        
        if (int(compareChoice2) == 2):
            print("Which variant for the Ford",evFordModel)
            print("1.",evFordVariant1)
            print("2.",evFordVariant2)
            print("")
            variantChoice = input("")
            print("")

            if (int(variantChoice) == 1):
                chosenEVMake = evFord

                chosenEVModel = evFordModel

                chosenEVPrice = evFordPrice1

                chosenEVVariant = evFordVariant1

                chosenEVBatteryCap = evFordBatteryCap1

                chosenEVElectricRange = evFordElectricRange1

                chosenEVCostToFullyCharge = (chosenEVBatteryCap * canadaAverageElectricityPrice) / 100

                chosenEVVehicleRunningCost = (chosenEVCostToFullyCharge / chosenEVElectricRange) * 100

            if (int(variantChoice) == 2):
                chosenEVMake = evFord

                chosenEVModel = evFordModel

                chosenEVPrice = evFordPrice2

                chosenEVVariant = evFordVariant2

                chosenEVBatteryCap = evFordBatteryCap2

                chosenEVElectricRange = evFordElectricRange2

                chosenEVCostToFullyCharge = (chosenEVBatteryCap * canadaAverageElectricityPrice) / 100

                chosenEVVehicleRunningCost = (chosenEVCostToFullyCharge / chosenEVElectricRange) * 100

        if (int(compareChoice2) == 3):
            chosenEVMake = rivian

            chosenEVModel = rivianModel

            chosenEVPrice = rivianPrice

            chosenEVVariant = rivianVariant

            chosenEVBatteryCap = rivianBatteryCap

            chosenEVElectricRange = rivianElectricRange

            chosenEVCostToFullyCharge = (chosenEVBatteryCap * canadaAverageElectricityPrice) / 100

            chosenEVVehicleRunningCost = (chosenEVCostToFullyCharge / chosenEVElectricRange) * 100
        
        print("You have chosen:",chosenEVMake,chosenEVModel)
        print("That come with a",chosenEVVariant,"variant")
        print("")

        print("1. Ford",fordModel)
        print("2. GMC",gmcModel)
        print("3. Chevrolet",chevModel)
        print("")
        compareChoice1 = input("Now pick an ICE Truck to compare it with: ")
        print("")

        if (int(compareChoice1) == 1):
            print("Which engine for the Ford",fordModel)
            print("1.",fordEngine1)
            print("2.",fordEngine2)
            print("3.",fordEngine3)
            print("")
            engineChoice = input("")
            print("")

            if (int(engineChoice) == 1):
                chosenICEMake = ford

                chosenICEModel = fordModel

                chosenICEPrice = fordPrice1

                chosenICEEngine = fordEngine1

                chosenICEFuelCap = fordFuelCap1

                chosenICEFuel = fordFuel1

                chosenICECityMilage = fordCityMilage1

                chosenICEHighwayMilage = fordHighwayMilage1

                chosenICEFullTank = (chosenICEFuelCap * canadaAverageGasPrice) / 100

                chosenICECityVehicleRunningCost = canadaAverageGasPrice / chosenICECityMilage

                chosenICEHighwayVehicleRunningCost = canadaAverageGasPrice / chosenICEHighwayMilage
            
            if (int(engineChoice) == 2):
                chosenICEMake = ford

                chosenICEModel = fordModel

                chosenICEPrice = fordPrice2

                chosenICEEngine = fordEngine2

                chosenICEFuelCap = fordFuelCap2

                chosenICEFuel = fordFuel2

                chosenICECityMilage = fordCityMilage2

                chosenICEHighwayMilage = fordHighwayMilage2

                chosenICEFullTank = (chosenICEFuelCap * canadaAverageGasPrice) / 100

                chosenICECityVehicleRunningCost = canadaAverageGasPrice / chosenICECityMilage

                chosenICEHighwayVehicleRunningCost = canadaAverageGasPrice / chosenICEHighwayMilage
            
            if (int(engineChoice) == 3):
                chosenICEMake = ford

                chosenICEModel = fordModel

                chosenICEPrice = fordPrice3

                chosenICEEngine = fordEngine3

                chosenICEFuelCap = fordFuelCap3

                chosenICEFuel = fordFuel3

                chosenICECityMilage = fordCityMilage3

                chosenICEHighwayMilage = fordHighwayMilage3

                chosenICEFullTank = (chosenICEFuelCap * canadaAverageGasPrice) / 100

                chosenICECityVehicleRunningCost = canadaAverageGasPrice / chosenICECityMilage

                chosenICEHighwayVehicleRunningCost = canadaAverageGasPrice / chosenICEHighwayMilage
        
        if (int(compareChoice1) == 2):
            chosenICEMake = gmc

            chosenICEModel = gmcModel

            chosenICEPrice = gmcPrice

            chosenICEEngine = gmcEngine

            chosenICEFuelCap = gmcFuelCap

            chosenICEFuel = gmcFuel

            chosenICECityMilage = gmcCityMilage

            chosenICEHighwayMilage = gmcHighwayMilage

            chosenICEFullTank = (chosenICEFuelCap * canadaAverageGasPrice) / 100

            chosenICECityVehicleRunningCost = canadaAverageGasPrice / chosenICECityMilage

            chosenICEHighwayVehicleRunningCost = canadaAverageGasPrice / chosenICEHighwayMilage
        
        if (int(compareChoice1) == 3):
            chosenICEMake = chev

            chosenICEModel = chevModel

            chosenICEPrice = chevPrice

            chosenICEEngine = chevEngine

            chosenICEFuelCap = chevFuelCap

            chosenICEFuel = chevFuel

            chosenICECityMilage = chevCityMilage

            chosenICEHighwayMilage = chevHighwayMilage

            chosenICEFullTank = (chosenICEFuelCap * canadaAverageGasPrice) / 100

            chosenICECityVehicleRunningCost = canadaAverageGasPrice / chosenICECityMilage

            chosenICEHighwayVehicleRunningCost = canadaAverageGasPrice / chosenICEHighwayMilage
        
        print("You have chosen:",chosenICEMake,chosenICEModel)
        print("That come with a",chosenICEEngine,"engine")
        print("")
        fleetCount = input("What is the fleet count? ")
        print("")
        input("Press enter to start comparing")
        print("")
        print("-----------------------------------------------------")
        print("Make:",chosenEVMake,"|",chosenICEMake)
        print("")
        print("Model:",chosenEVModel,"|",chosenICEModel)
        print("")
        print("Base Price:",chosenEVPrice,"|",chosenICEPrice)
        print("")
        fleetCountEVBasePrice = float(fleetCount) * chosenEVPrice
        fleetCountICEBasePrice = float(fleetCount) * chosenICEPrice
        print("Price based on fleet count:",f'{fleetCountEVBasePrice:.2f}',"|",f'{fleetCountICEBasePrice:.2f}')
        print("")
        priceDifference = chosenEVPrice - chosenICEPrice
        print("The base price difference between this EV truck")
        print("to ICE truck is:",f'{priceDifference:.2f}')
        print("")
        fleetCountPriceDifference = priceDifference * float(fleetCount)
        print("Price difference based on fleet count:",f'{fleetCountPriceDifference:.2f}')
        print("")
        if chosenEVPrice < 60000:
            print("Since the base price of this EV is under 60000,")
            print("this EV is eligible for a federal EV rebate of 5000.")
            print("Plus possible additional EV rebate depending on your province")
        print("")
        print("Cost for full tank ($):",f'{chosenICEFullTank:.2f}')
        print("Cost for full battery ($):",f'{chosenEVCostToFullyCharge:.2f}')
        print("")
        print("The average person fills their car's once a week")
        costToFillTankAnnually = chosenICEFullTank * 52
        print(f'{chosenICEFullTank:.2f}',"X 52 weeks =",f'{costToFillTankAnnually:.2f}')
        print("")
        fleetCountAnnualFill = costToFillTankAnnually * float(fleetCount)
        print("Cost to fill tank annually based on fleet count:",f'{fleetCountAnnualFill:.2f}')
        print("")
        print("It's been recommended to charge an EV about every 3 days.")
        print("Based on the recommendation, an EV will charge 2 times a week")
        costToChargeAnnually = (chosenEVCostToFullyCharge * 2) * 52
        print("(",f'{chosenEVCostToFullyCharge:.2f}',"X 2) X 52 Weeks =",f'{costToChargeAnnually:.2f}')
        print("")
        fleetCountAnnualCharge = costToChargeAnnually * float(fleetCount)
        print("Cost of charging battery based on fleet count:",f'{fleetCountAnnualCharge:.2f}')
        print("")
        annualPriceDifference = costToFillTankAnnually - costToChargeAnnually
        print("The annual cost difference between a full tank to a full battery:",f'{annualPriceDifference:.2f}')
        fleetCountAnnualPriceDifference = annualPriceDifference * float(fleetCount)
        print("With fleet count:",f'{fleetCountAnnualPriceDifference:.2f}')
        print("")
        print(chosenICEMake,chosenICEModel,"Vehicle running cost (¢/km)")
        print("City:",f'{chosenICECityVehicleRunningCost:.2f}',"Highway:",f'{chosenICEHighwayVehicleRunningCost:.2f}')
        fleetCountCityVehicleRunningCost = chosenICECityVehicleRunningCost * float(fleetCount)
        fleetCountHighwayVehicleRunningCost = chosenICEHighwayVehicleRunningCost * float(fleetCount)
        print("With fleet count")
        print("City:",f'{fleetCountCityVehicleRunningCost:.2f}',"Highway:",f'{fleetCountHighwayVehicleRunningCost:.2f}')
        print("")
        print(chosenEVMake,chosenEVModel,"Vehicle running cost (¢/km):",f'{chosenEVVehicleRunningCost:.2f}')
        fleetCountEVVehicleRunningCost = chosenEVVehicleRunningCost * float(fleetCount)
        print("With fleet count:",f'{fleetCountEVVehicleRunningCost:.2f}')
        print("-----------------------------------------------------")
        input("Press enter to see cost over time")
        print("")
        totalICECost = chosenICEPrice + costToFillTankAnnually
        fleetCountTotalICECost = totalICECost * float(fleetCount)
        print("Total cost after 1 year of having ICE Truck:",f'{totalICECost:.2f}')
        print("With fleet count:", f'{fleetCountTotalICECost:.2f}')
        print("")
        totalEVCost = chosenEVPrice + costToChargeAnnually
        fleetCountTotalEVCost = totalEVCost * float(fleetCount)
        print("Total cost after 1 year of having EV Truck:",f'{totalEVCost:.2f}')
        print("With fleet count:", f'{fleetCountTotalEVCost:.2f}')
        print("")
        print("")
        totalICECost3 = (costToFillTankAnnually * 3) + chosenICEPrice
        fleetCountTotalICECost3 = totalICECost3 * float(fleetCount)
        print("Total cost after 3 years of having ICE Truck:",f'{totalICECost3:.2f}')
        print("With fleet count:", f'{fleetCountTotalICECost3:.2f}')
        print("")
        totalEVCost3 = (costToChargeAnnually * 3) + chosenEVPrice
        fleetCountTotalEVCost3 = totalEVCost3 * float(fleetCount)
        print("Total cost after 3 years of having EV Truck:",f'{totalEVCost3:.2f}')
        print("With fleet count:", f'{fleetCountTotalEVCost3:.2f}')
        print("")
        print("")
        totalICECost5 = (costToFillTankAnnually * 5) + chosenICEPrice
        fleetCountTotalICECost5 = totalICECost5 * float(fleetCount)
        print("Total cost after 5 years of having ICE Truck:",f'{totalICECost5:.2f}')
        print("With fleet count:", f'{fleetCountTotalICECost5:.2f}')
        print("")
        totalEVCost5 = (costToChargeAnnually * 5) + chosenEVPrice
        fleetCountTotalEVCost5 = totalEVCost5 * float(fleetCount)
        print("Total cost after 5 years of having EV Truck:",f'{totalEVCost5:.2f}')
        print("With fleet count:", f'{fleetCountTotalEVCost5:.2f}')
        print("-----------------------------------------------------")

        input("Press enter to return to main menu")

    else:
        print("That is not a valid choice, please input a number from 1 to 4")
        input("Press enter to continue")

